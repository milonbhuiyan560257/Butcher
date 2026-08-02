from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os


class ExcelManager:

    def __init__(self, filename="voucher.xlsx"):
        self.filename = filename
        self._ensure_file()

    def _ensure_file(self):
        if not os.path.exists(self.filename):
            wb = Workbook()
            ws = wb.active
            ws.title = "Voucher"
            
            headers = ["Name", "Voucher No", "Amount", "Mobile", "Date", "Time"]
            ws.append(headers)
            
            # Header styling
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=12)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            
            # Auto-adjust column widths
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 12
            
            wb.save(self.filename)

    def save_data(self, name, voucher, amount, mobile, date, time):
        self._ensure_file()
        wb = load_workbook(self.filename)
        ws = wb.active
        
        row_data = [name, voucher, amount, mobile, date, time]
        ws.append(row_data)
        
        # Style the new row
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        for cell in ws[ws.max_row]:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        wb.save(self.filename)

    def total_rows(self):
        if not os.path.exists(self.filename):
            return 0
        wb = load_workbook(self.filename)
        ws = wb.active
        return max(0, ws.max_row - 1)
    
    def get_all_data(self):
        if not os.path.exists(self.filename):
            return []
        wb = load_workbook(self.filename)
        ws = wb.active
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            data.append(row)
        return data
