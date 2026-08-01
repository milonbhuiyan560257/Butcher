from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.button import Button
from kivy.uix.label import Label
from kivy.uix.camera import Camera
from kivy.uix.textinput import TextInput
from kivy.clock import Clock

import os
from openpyxl import Workbook, load_workbook


class VoucherScanner(BoxLayout):

    def __init__(self, **kwargs):
        super().__init__(orientation="vertical", **kwargs)

        self.camera = Camera(
            play=True,
            resolution=(1280, 720)
        )

        self.add_widget(self.camera)

        self.status = Label(
            text="Ready",
            size_hint=(1, 0.08)
        )

        self.add_widget(self.status)

        self.name = TextInput(
            hint_text="Child Name",
            multiline=False,
            size_hint=(1, 0.08)
        )

        self.add_widget(self.name)

        self.voucher = TextInput(
            hint_text="Voucher Number",
            multiline=False,
            size_hint=(1, 0.08)
        )

        self.add_widget(self.voucher)

        self.amount = TextInput(
            hint_text="Amount",
            multiline=False,
            size_hint=(1, 0.08)
        )

        self.add_widget(self.amount)

        btn1 = Button(
            text="Capture Image",
            size_hint=(1, 0.09)
        )
        btn1.bind(on_press=self.capture)

        self.add_widget(btn1)

        btn2 = Button(
            text="Save Excel",
            size_hint=(1, 0.09)
        )
        btn2.bind(on_press=self.save_excel)

        self.add_widget(btn2)

        btn3 = Button(
            text="Start OCR (Next Step)",
            size_hint=(1, 0.09)
        )
        btn3.bind(on_press=self.start_ocr)

        self.add_widget(btn3)

    def capture(self, instance):

        if not os.path.exists("captures"):
            os.mkdir("captures")

        filename = "captures/voucher.png"

        self.camera.export_to_png(filename)

        self.status.text = "Image Saved"

    def save_excel(self, instance):

        file = "voucher.xlsx"

        if os.path.exists(file):
            wb = load_workbook(file)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append([
                "Name",
                "Voucher",
                "Amount"
            ])

        ws.append([
            self.name.text,
            self.voucher.text,
            self.amount.text
        ])

        wb.save(file)

        self.status.text = "Saved To Excel"

    def start_ocr(self, instance):

        self.status.text = "OCR Module Coming Next"


class VoucherApp(App):

    def build(self):

        return VoucherScanner()


if __name__ == "__main__":
    VoucherApp().run()
